import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter
from typing import List, Dict
import chromadb
from sentence_transformers import SentenceTransformer
from llama_cpp import Llama
import os
import json
import hashlib

class ExcelEmbeddingProcessor:
    def __init__(self, db_path: str = "./chroma_db", cache_dir: str = "./embedding_cache"):
        """
        Initialize the processor with paths for ChromaDB and embedding cache.
        
        Args:
            db_path (str): Path to ChromaDB storage
            cache_dir (str): Path to store embedding cache files
        """
        self.client = chromadb.PersistentClient(path=db_path)
        self.cache_dir = cache_dir
        os.makedirs(cache_dir, exist_ok=True)
        self.embedding_model = SentenceTransformer("all-MiniLM-L6-v2")

    def _get_file_hash(self, file_path: str) -> str:
        """
        Generate a hash of the Excel file content for cache identification.
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            str: Hash of the file content
        """
        with open(file_path, 'rb') as f:
            file_content = f.read()
            return hashlib.md5(file_content).hexdigest()

    def _get_cache_path(self, file_hash: str) -> str:
        """
        Get the path where cached embeddings should be stored.
        
        Args:
            file_hash (str): Hash of the Excel file
            
        Returns:
            str: Path to the cache file
        """
        return os.path.join(self.cache_dir, f"{file_hash}_embeddings.json")

    def _cache_exists(self, file_hash: str) -> bool:
        """
        Check if cached embeddings exist for the given file hash.
        
        Args:
            file_hash (str): Hash of the Excel file
            
        Returns:
            bool: True if cache exists, False otherwise
        """
        cache_path = self._get_cache_path(file_hash)
        return os.path.exists(cache_path)

    def process_excel_file(self, file_path: str) -> List[Dict]:
        """
        Process an Excel file to generate chunks with headers and merged cell hierarchy.
        [Previous implementation remains the same]
        """
        # [Previous implementation of process_excel_file remains unchanged]
        wb = openpyxl.load_workbook(file_path)
        chunks = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Extract headers from the first row
            headers = {}
            header_row = next(sheet.iter_rows(min_row=1, max_row=1))
            for cell in header_row:
                header_value = cell.value
                if header_value is None:
                    header_value = get_column_letter(cell.column)
                headers[cell.column] = header_value

            # Identify vertical merges
            vertical_merges = []
            for merged_range in sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(merged_range.coord)
                if min_col == max_col and min_row < max_row and min_row >= 2:
                    parent_value = sheet.cell(row=min_row, column=min_col).value
                    vertical_merges.append({
                        "min_col": min_col,
                        "min_row": min_row,
                        "max_row": max_row,
                        "parent_value": parent_value
                    })

            # Process data rows
            processed_rows = []
            for row in sheet.iter_rows(min_row=2):
                row_data = {}
                for cell in row:
                    cell_value = cell.value
                    for vm in vertical_merges:
                        if vm["min_col"] == cell.column and vm["min_row"] <= cell.row <= vm["max_row"]:
                            cell_value = vm["parent_value"]
                            break
                    header = headers.get(cell.column, get_column_letter(cell.column))
                    row_data[header] = cell_value
                processed_rows.append(row_data)

            # Generate chunks
            for vm in vertical_merges:
                start_idx = vm["min_row"] - 2
                end_idx = vm["max_row"] - 2
                chunk_rows = processed_rows[start_idx: end_idx + 1]

                chunk_text = ""
                for row in chunk_rows:
                    row_text = ", ".join([f"{k}: {v}" for k, v in row.items() if v is not None])
                    chunk_text += f"{row_text}\n"
                chunk_text = chunk_text.strip()

                chunks.append({
                    "sheet": sheet_name,
                    "parent_value": vm["parent_value"],
                    "chunk_text": chunk_text
                })

        return chunks

    def create_embeddings(self, file_path: str, collection_name: str = "excel_data") -> None:
        """
        Create embeddings for the Excel file chunks and store them in ChromaDB and cache.
        
        Args:
            file_path (str): Path to the Excel file
            collection_name (str): Name of the ChromaDB collection
        """
        file_hash = self._get_file_hash(file_path)
        collection = self.client.get_or_create_collection(name=collection_name)

        # Check if embeddings already exist in cache
        if self._cache_exists(file_hash):
            print("Loading embeddings from cache...")
            cache_path = self._get_cache_path(file_hash)
            with open(cache_path, 'r') as f:
                cached_data = json.load(f)

            # Add cached embeddings to ChromaDB if not already present
            if not self._embeddings_exist_in_chroma(collection, file_path):
                print("Adding cached embeddings to ChromaDB...")
                collection.add(
                    ids=cached_data["ids"],
                    embeddings=cached_data["embeddings"],
                    metadatas=cached_data["metadatas"]
                )
        else:
            print("Creating new embeddings...")
            chunks = self.process_excel_file(file_path)
            
            # Prepare data for both ChromaDB and cache
            embeddings_data = {
                "ids": [],
                "embeddings": [],
                "metadatas": []
            }

            for idx, chunk in enumerate(chunks):
                embedding = self.embedding_model.encode(chunk["chunk_text"]).tolist()
                metadata = {
                    "sheet": chunk["sheet"],
                    "parent_value": chunk["parent_value"],
                    "chunk_text": chunk["chunk_text"],
                    "source": file_path
                }
                
                embeddings_data["ids"].append(str(idx))
                embeddings_data["embeddings"].append(embedding)
                embeddings_data["metadatas"].append(metadata)

            # Store in ChromaDB
            collection.add(
                ids=embeddings_data["ids"],
                embeddings=embeddings_data["embeddings"],
                metadatas=embeddings_data["metadatas"]
            )

            # Cache the embeddings
            cache_path = self._get_cache_path(file_hash)
            with open(cache_path, 'w') as f:
                json.dump(embeddings_data, f)

            print(f"Successfully processed {len(chunks)} chunks and cached embeddings.")

    def _embeddings_exist_in_chroma(self, collection, file_path: str) -> bool:
        """
        Check if embeddings exist in ChromaDB.
        
        Args:
            collection: ChromaDB collection
            file_path (str): Path to the Excel file
            
        Returns:
            bool: True if embeddings exist, False otherwise
        """
        existing = collection.get(where={"source": file_path})
        return len(existing["ids"]) > 0

    def query_collection(self, query_text: str, collection_name: str = "excel_data", top_k: int = 3):
        """
        Query the ChromaDB collection for similar chunks.
        
        Args:
            query_text (str): Query text to search for
            collection_name (str): Name of the ChromaDB collection
            top_k (int): Number of top results to return
        """
        collection = self.client.get_or_create_collection(name=collection_name)
        query_embedding = self.embedding_model.encode(query_text).tolist()

        results = collection.query(
            query_embeddings=[query_embedding],
            n_results=top_k
        )

        # print(results)

        # print("\nQuery results:")
        # for i, (chunk_id, metadata) in enumerate(zip(results["ids"][0], results["metadatas"][0])):
        #     print(f"\nResult {i + 1}:")
        #     print(f"Sheet: {metadata['sheet']}")
        #     print(f"Parent Value: {metadata['parent_value']}")
        #     print(f"Chunk Text: {metadata['chunk_text']}")
        #     print("-" * 50)
        return results
        
    def augment_user_query(self, query_text: str, top_k: int = 5) -> str:
        """
        Augment the user query with relevant context from the Excel data.
        
        Args:
            query_text (str): User query text.
            top_k (int): Number of similar chunks to retrieve.
            
        Returns:
            str: Augmented query text with relevant context.
        """
        relevant_chunks = self.query_collection(query_text, top_k=top_k)
       
        augmented_query = f"""
          ### User Scenario:
          {query_text}

          


          You are an expert in threat and vulnerability analysis. Given the scenario above, please provide:
          1. **Predicted Remediation Strategy(ies)**
          2. **Reasoning**
          3. **Classification Description**
          4. **Threat ID**
          5. **Vulnerability ID**
          6. **Remediation ID (Countermeasure ID)**

          based on the following Relevant Context:
          {relevant_chunks}
              
          Format your answer exactly as specified.
              """
        return augmented_query.strip()

    def generate_response(self, prompt):
        qwens = Llama(
            model_path="models/llama-o1-supervised-1129-q4_k_m.gguf",
            n_ctx=4096,
            chat_format="qwen",
            main_gpu=0
        )
        response = qwens.create_chat_completion(
            messages=[
                {"role": "system", "content": "You are an AI assistant helping with risk analysis of threats and vulnerabilities in a mission-critical system. Provide responses based on the given context without fabricating information. Your response should be well formatted, including : threat ID, threat name, description, vulnerability ID, name, description, countermeasure ID, and your reasoning."},
                {"role": "user", "content": prompt}
            ],
            stream=True
        )

        for chunk in response:
            print(chunk['choices'][0]['delta'].get('content', ''), end='', flush=True)

def main():
    processor = ExcelEmbeddingProcessor()
    excel_path = "test_data/rb.xlsx"
    
    # Create/load embeddings
    processor.create_embeddings(excel_path)
    
    # Example query
    query_text = "Confidential documents are stored in an archive constantly protected by armed guards, with three levels of biometric protection. The room that houses the archive is reinforced and burglar-proof."
    augmented_query = processor.augment_user_query(query_text, top_k=1)
    print(augmented_query)
    processor.generate_response(augmented_query)

if __name__ == "__main__":
    main()